import shutil as sh
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

class Gerador_de_desligamentos:
    def __init__(self):
        self.__MRE = pd.read_excel(os.path.join("utils/gerador_de_desligamentos/dados","Mre.xlsx"))
        self.__SCE = pd.read_excel(os.path.join("utils/gerador_de_desligamentos/dados","Sce.xlsx"))
        self.__EXIT_PATH = os.path.join(os.path.join(os.path.expanduser("~"), 'Downloads'), f"Desligamentos {datetime.now().strftime('%Y_%m_%d')}")
        self.__MODELO = os.path.join("utils/gerador_de_desligamentos/dados",'Modelo_desligamento.xlsx')
        self.__RECESSOS = pd.read_excel(os.path.join("utils/gerador_de_desligamentos/dados","Recessos.xlsx"))
        self.__FALTAS = pd.read_excel(os.path.join("utils/gerador_de_desligamentos/dados","Faltas.xlsx"))

        self.__gerar_pasta_saida()

        self.__desligados = {
            'Nome' : [],
            'CPF' : [],
            'DataInicio' : [],
            'DataFinal' : [],
            'Bolsa' : []
        }

        self.__mre_cpf = []

    def __gerar_pasta_saida(self):
        if not os.path.isdir(self.__EXIT_PATH):
            os.makedirs(self.__EXIT_PATH)

    def __conversor_de_cpf(self, cpf):
        if len(cpf) != 11 and cpf[0] != '0' and '.' not in cpf:
            cpf = '0' + cpf
            if len(cpf) != 11:
                pass
            else:
                return str(cpf)
        elif len(cpf) != 11:
                return str(cpf[:3] + cpf[4:7] + cpf[8:11] + cpf[12:])
        else:
            return str(cpf)

    def __gerar_dados(self, cpf, data_alternativa):
        for a in range(len(self.__MRE.iloc[:,3])):
            self.__mre_cpf.append((self.__conversor_de_cpf(str(self.__MRE.iloc[a,4]))))

        for b in range(len(self.__SCE.iloc[:,6])):
                if self.__conversor_de_cpf(str(cpf)) == self.__conversor_de_cpf(str(self.__SCE.iloc[b,6])):
                    self.__desligados['Nome'].append(str(self.__SCE.iloc[b,4]))
                    self.__desligados['CPF'].append(self.__conversor_de_cpf(str(self.__SCE.iloc[b,6])))
                    self.__desligados['DataInicio'].append(str(self.__SCE.iloc[b,23]).split(' a ')[0])
                    self.__desligados['DataFinal'].append(str(self.__SCE.iloc[b,23]).split(' a ')[-1])
                    self.__desligados['Bolsa'].append(str(self.__SCE.iloc[b,22]))

        for c in self.__desligados['CPF']:
            if c in self.__mre_cpf:
                for d in range(len(self.__MRE.iloc[:,4])):
                    if c == self.__conversor_de_cpf(str(self.__MRE.iloc[d,4])):
                        sh.copy(self.__MODELO, os.path.join(self.__EXIT_PATH, f"{self.__desligados['Nome'][self.__desligados['CPF'].index(c)]}.xlsx"))
                        aux = load_workbook(os.path.join(self.__EXIT_PATH, f"{self.__desligados['Nome'][self.__desligados['CPF'].index(c)]}.xlsx"))
                        aba = aux.active
                        aba['B2'] = f"{self.__desligados['Nome'][self.__desligados['CPF'].index(c)]}" 
                        aba['B3'] = c
                        aba['B4'] = f"{str(self.__MRE.iloc[d,2])}"
                        aba['E5'] = f"{self.__MRE.iloc[d,26].strftime('%d/%m/%Y')}"
                        if data_alternativa == 'Insira data de deligamento alternativa' or data_alternativa == '':
                            aba['E6'] = f"{(self.__MRE.iloc[d,27].strftime('%d/%m/%Y'))}"
                            if 30 - int((self.__MRE.iloc[d,27].strftime("%d"))) <= 1:
                                aba['C21'] = 0
                                aba['C28'] = 0
                            else:
                                aba['C21'] = 30 - int((self.__MRE.iloc[d,27].strftime("%d")))
                                aba['C28'] = 30 - int((self.__MRE.iloc[d,27].strftime("%d")))
                        else:
                            aba['E6'] = data_alternativa
                            if 30 - int(data_alternativa[:2]) <= 1:
                                aba['C21'] = 0
                                aba['C28'] = 0
                            else:
                                aba['C21'] = 30 - int(data_alternativa[:2])
                                aba['C28'] = 30 - int(data_alternativa[:2])
                        if str(self.__MRE.iloc[d,17]) == 30:
                            aba['B7'] = '1125,69'
                        else:
                            aba['B7'] = '787,98'
                        aba['C10'] = f"{str(self.__RECESSOS.iloc[0,2])}"
                        for v in range(len(self.__FALTAS.iloc[:,0])):
                            # 037.991.061-62
                            if str(self.__FALTAS.iloc[v,0].split(' - ')[0]) == str(self.__MRE.iloc[d,2]):
                                if int(self.__FALTAS.iloc[v,1]) > 0:
                                    aba['C27'] = int(self.__FALTAS.iloc[v,1])/10
                                    if float(self.__FALTAS.iloc[v,2]) > 0:
                                        if str(self.__MRE.iloc[d,17]) == 30:
                                            aba['C22'] = int(self.__FALTAS.iloc[v,2])//(1125.69/30)-10
                                            aba['C27'] = int(self.__FALTAS.iloc[v,1])/10 + (int(self.__FALTAS.iloc[v,2])//(1125.69/30)-10)
                                            break
                                        else:
                                            aba['C22'] = int(self.__FALTAS.iloc[v,2])//(787.98/30)-10
                                            aba['C27'] = int(self.__FALTAS.iloc[v,1])/10 + (int(self.__FALTAS.iloc[v,2])//(787.98/30)-10)
                                            break
                                    else:
                                        aba['C22'] = 0
                                        break
                                else:
                                    aba['C27'] = 0
                                    if float(self.__FALTAS.iloc[v,2]) > 0:
                                        if str(self.__MRE.iloc[d,17]) == 30:
                                            aba['C22'] = int(self.__FALTAS.iloc[v,2])//(1125.69/30)-10
                                            aba['C27'] = int(self.__FALTAS.iloc[v,1])/10 + (int(self.__FALTAS.iloc[v,2])//(1125.69/30)-10)
                                            break
                                        else:
                                            aba['C22'] = int(self.__FALTAS.iloc[v,2])//(787.98/30)
                                            aba['C27'] = int(self.__FALTAS.iloc[v,1])/10 + (int(self.__FALTAS.iloc[v,2])//(787.98/30)-10)
                                            break
                                    else:
                                        aba['C22'] = 0
                                        break
                            else:
                                continue
                        aux.save(os.path.join(self.__EXIT_PATH, f"{self.__desligados['Nome'][self.__desligados['CPF'].index(c)]}.xlsx"))

            else:
                sh.copy(self.__MODELO, os.path.join(self.__EXIT_PATH, f"{self.__desligados['Nome'][self.__desligados['CPF'].index(c)]}.xlsx"))
                aux = load_workbook(os.path.join(self.__EXIT_PATH, f"{self.__desligados['Nome'][self.__desligados['CPF'].index(c)]}.xlsx"))
                aba = aux.active
                aba['B2'] = f"{self.__desligados['Nome'][self.__desligados['CPF'].index(c)]}" 
                aba['B3'] = c
                aba['B4'] = f"XXX"
                aba['B7'] = f"{self.__desligados['Bolsa'][self.__desligados['CPF'].index(c)]}"
                aba['E5'] = f"{self.__desligados['DataInicio'][self.__desligados['CPF'].index(c)]}"
                if data_alternativa == 'Insira data de deligamento alternativa' or data_alternativa == '':
                    aba['E6'] = f"{self.__desligados['DataFinal'][self.__desligados['CPF'].index(c)]}"
                    if 30 - int(self.__desligados['DataFinal'][self.__desligados['CPF'].index(c)][:2]) <= 1:
                        aba['C21'] = 0
                    else:
                        aba['C21'] = 30 - int(self.__desligados['DataFinal'][self.__desligados['CPF'].index(c)][:2])
                else:
                    aba['E6'] = data_alternativa
                    if 30 - int(data_alternativa[:2]) <= 1:
                        aba['C21'] = 0
                    else:
                        aba['C21'] = 30 - int(data_alternativa[:2])
                aba['C10'] = f"{str(self.__RECESSOS.iloc[0,2])}"
                aux.save(os.path.join(self.__EXIT_PATH, f"{self.__desligados['Nome'][self.__desligados['CPF'].index(c)]}.xlsx"))
                
    def __limpar_listas(self):
        for chave in self.__desligados:
            self.__desligados[chave].clear()

        self.__mre_cpf.clear()
    
    def iniciar(self, cpf, data_alternativa):    
        self.__gerar_dados(cpf, data_alternativa)
        self.__limpar_listas()